"""Bulk shipment import from carrier Excel / CSV exports.

Carriers (Viettel Post, GHN, GHTK, ...) hand operators monthly delivery
dumps. Each carrier picks its own column names, encodings, and header
offsets. This module reads those files, maps columns via a per-carrier
profile (``config.features.shipment.import.profiles.<name>``), matches each
row's recipient phone against the students table to resolve SBDs, and
upserts shipment rows.

History semantics
-----------------

Shipment PK is ``(round_id, sbd, tracking_code)``. A re-send attempt for the
same SBD produces a new row keyed by the new tracking code — the audit
trail keeps every carrier attempt for a student. Existing rows get
overwritten on ``INSERT OR REPLACE`` only when the *same* tracking code
arrives with updated status (carrier delivery-in-progress → delivered).

Safety defaults
---------------

The handler never writes unless ``commit=True``. Callers default to a
dry-run that returns :class:`BulkImportStats` so operators can review the
outcome before persisting.
"""

from __future__ import annotations

import csv
import re
import sqlite3
from collections import Counter
from contextlib import closing
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from luonvuitoi_cert.auth.activity_log import ActivityLog, log_admin_action
from luonvuitoi_cert.config import CertConfig
from luonvuitoi_cert.config.models import ShipmentImportProfile

_CREATE_SQL = """
CREATE TABLE IF NOT EXISTS shipment_history (
    round_id TEXT NOT NULL,
    sbd TEXT NOT NULL,
    tracking_code TEXT NOT NULL,
    is_success INTEGER NOT NULL DEFAULT 0,
    status_raw TEXT,
    sent_at TEXT,
    phone TEXT,
    address TEXT,
    recipient TEXT,
    carrier TEXT NOT NULL,
    synced_at TEXT,
    PRIMARY KEY (round_id, sbd, tracking_code)
);
CREATE INDEX IF NOT EXISTS idx_shipment_history_sbd ON shipment_history(sbd);
CREATE INDEX IF NOT EXISTS idx_shipment_history_tracking ON shipment_history(tracking_code);
"""


class BulkImportError(Exception):
    """Raised for missing config, unknown carrier, unreadable files, etc."""


@dataclass(frozen=True, slots=True)
class BulkImportStats:
    carrier: str
    round_id: str
    parsed: int = 0
    skipped_prefix: int = 0
    skipped_no_tracking: int = 0
    skipped_no_phone: int = 0
    matched_sbds: int = 0
    unmatched_phones: int = 0
    inserted: int = 0
    success_count: int = 0
    committed: bool = False
    status_breakdown: dict[str, int] = field(default_factory=dict)


# --- helpers ----------------------------------------------------------------


def _normalize_phone(raw: Any) -> str:
    """Strip non-digits + leading zero (VN convention).

    Carrier sheets store phones as floats (``901234567.0``), strings with
    leading 0 (``0901234567``), or with separators (``+84 90 123 4567``).
    Normalize to bare digits, no leading 0, so lookup against students table
    is consistent regardless of source format.
    """
    if raw is None:
        return ""
    s = str(raw).strip()
    if s.endswith(".0"):
        s = s[:-2]
    digits = re.sub(r"\D", "", s)
    return digits.lstrip("0")


def _cell_to_str(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _resolve_carrier(config: CertConfig, carrier: str | None) -> tuple[str, ShipmentImportProfile]:
    if config.features.shipment.import_ is None:
        raise BulkImportError(
            "shipment.import block missing from config — add profiles before running bulk import"
        )
    imp = config.features.shipment.import_
    name = carrier or imp.default
    if not name:
        raise BulkImportError(f"no --carrier given and no default set; available: {sorted(imp.profiles)}")
    if name not in imp.profiles:
        raise BulkImportError(f"unknown carrier {name!r}; available: {sorted(imp.profiles)}")
    return name, imp.profiles[name]


def _require_phone_col(config: CertConfig) -> str:
    col = config.data_mapping.phone_col
    if not col:
        raise BulkImportError("data_mapping.phone_col must be set — bulk import matches SBD via phone")
    return col


def _require_round(config: CertConfig, round_id: str) -> str:
    for r in config.rounds:
        if r.id == round_id:
            return r.table
    raise BulkImportError(f"unknown round_id {round_id!r}; available: {[r.id for r in config.rounds]}")


def _first_matching_header(headers: list[str], fallbacks: list[str]) -> str | None:
    """Resolve the first header in ``headers`` that appears in ``fallbacks``.

    Case- and whitespace-tolerant so minor carrier header tweaks don't break
    mapping.
    """
    norm_headers = {h.strip().lower(): h for h in headers if h}
    for fb in fallbacks:
        h = norm_headers.get(fb.strip().lower())
        if h is not None:
            return h
    return None


def _read_xlsx(path: Path, header_row: int) -> tuple[list[str], list[dict[str, str]]]:
    try:
        wb = load_workbook(path, read_only=True, data_only=True)
    except Exception as e:
        raise BulkImportError(f"failed to open Excel workbook: {e}") from e
    try:
        ws = wb.active
        if ws is None:
            raise BulkImportError("workbook has no active worksheet")
        rows_iter = ws.iter_rows(values_only=True)
        # Skip pre-header metadata rows
        for _ in range(header_row):
            try:
                next(rows_iter)
            except StopIteration:
                return [], []
        try:
            header_raw = next(rows_iter)
        except StopIteration:
            return [], []
        headers = [_cell_to_str(h) for h in header_raw]
        rows: list[dict[str, str]] = []
        for row in rows_iter:
            values = [_cell_to_str(v) for v in row]
            values += [""] * (len(headers) - len(values))
            if all(v == "" for v in values[: len(headers)]):
                continue
            rows.append({h: v for h, v in zip(headers, values[: len(headers)], strict=False) if h})
        return [h for h in headers if h], rows
    finally:
        wb.close()


def _read_csv(path: Path, header_row: int) -> tuple[list[str], list[dict[str, str]]]:
    try:
        with path.open(encoding="utf-8-sig", newline="") as f:
            reader = csv.reader(f)
            for _ in range(header_row):
                next(reader, None)
            try:
                headers = [h.strip() for h in next(reader)]
            except StopIteration:
                return [], []
            rows: list[dict[str, str]] = []
            for row in reader:
                padded = list(row) + [""] * (len(headers) - len(row))
                rec = {h: padded[i].strip() for i, h in enumerate(headers) if h}
                if not any(rec.values()):
                    continue
                rows.append(rec)
            return headers, rows
    except OSError as e:
        raise BulkImportError(f"failed to read CSV: {e}") from e


def _read_input(path: Path, header_row: int) -> tuple[list[str], list[dict[str, str]]]:
    suffix = path.suffix.lower()
    if suffix in {".xlsx", ".xlsm"}:
        return _read_xlsx(path, header_row)
    if suffix == ".csv":
        return _read_csv(path, header_row)
    raise BulkImportError(f"unsupported file type {suffix!r}; use .xlsx, .xlsm, or .csv")


def _is_success(status: str, keywords: list[str]) -> bool:
    up = (status or "").strip().upper()
    return any(k.strip().upper() in up for k in keywords)


def _has_skip_prefix(status: str, prefixes: list[str]) -> bool:
    up = (status or "").strip().upper()
    return any(up.startswith(p.strip().upper()) for p in prefixes if p.strip())


def _build_phone_to_sbds(
    db_path: Path, round_table: str, sbd_col: str, phone_col: str
) -> dict[str, list[str]]:
    mapping: dict[str, list[str]] = {}
    with closing(sqlite3.connect(str(db_path))) as conn:
        conn.row_factory = sqlite3.Row
        try:
            rows = conn.execute(
                f'SELECT "{sbd_col}" AS sbd, "{phone_col}" AS phone FROM "{round_table}"'
            ).fetchall()
        except sqlite3.OperationalError as e:
            raise BulkImportError(f"cannot query students table {round_table!r}: {e}") from e
    for row in rows:
        phone = _normalize_phone(row["phone"])
        sbd = str(row["sbd"] or "").strip()
        if not phone or not sbd:
            continue
        mapping.setdefault(phone, []).append(sbd)
    return mapping


def _ensure_schema(conn: sqlite3.Connection) -> None:
    conn.executescript(_CREATE_SQL)


# --- public API -------------------------------------------------------------


def bulk_import_shipments(
    *,
    config: CertConfig,
    db_path: str | Path,
    activity: ActivityLog,
    file_path: str | Path,
    round_id: str,
    carrier: str | None = None,
    commit: bool = False,
    admin_user_id: str | None = None,
    admin_email: str | None = None,
    client_ip: str | None = None,
) -> BulkImportStats:
    """Parse a carrier export and (optionally) upsert its rows.

    Returns stats for the run. Does not mutate the DB when ``commit`` is
    false — callers display the stats to the operator before rerunning
    with ``commit=True``.
    """
    carrier_name, profile = _resolve_carrier(config, carrier)
    phone_col = _require_phone_col(config)
    round_table = _require_round(config, round_id)
    db_path = Path(db_path).expanduser().resolve()
    path = Path(file_path).expanduser().resolve()
    if not path.exists():
        raise BulkImportError(f"import file not found: {path}")

    headers, rows = _read_input(path, profile.header_row)
    mapping = profile.column_mapping

    col_tracking = _first_matching_header(headers, mapping.tracking_code)
    col_phone = _first_matching_header(headers, mapping.phone)
    col_status = _first_matching_header(headers, mapping.status)
    col_sent_at = _first_matching_header(headers, mapping.sent_at)
    col_address = _first_matching_header(headers, mapping.address)
    col_recipient = _first_matching_header(headers, mapping.recipient)

    missing = [
        logical
        for logical, resolved in (
            ("tracking_code", col_tracking),
            ("phone", col_phone),
            ("status", col_status),
        )
        if resolved is None
    ]
    if missing:
        raise BulkImportError(f"carrier {carrier_name!r}: headers {missing} unresolved; file has {headers!r}")

    phone_to_sbds = _build_phone_to_sbds(db_path, round_table, config.data_mapping.sbd_col, phone_col)

    # Dedup by tracking code — first row wins (stable file order).
    seen_tracking: set[str] = set()
    to_insert: list[tuple[str, str, str, int, str, str, str, str, str, str, str]] = []
    skipped_prefix = 0
    skipped_no_tracking = 0
    skipped_no_phone = 0
    unmatched_phones: set[str] = set()
    matched_sbds: set[str] = set()
    success_count = 0
    status_counter: Counter[str] = Counter()

    # synced_at: single timestamp for whole run
    synced_at = _iso_now()

    for row in rows:
        tracking = row.get(col_tracking, "").strip() if col_tracking else ""
        status = row.get(col_status, "").strip() if col_status else ""
        phone = _normalize_phone(row.get(col_phone, "") if col_phone else "")

        if not tracking:
            skipped_no_tracking += 1
            continue
        if tracking in seen_tracking:
            continue
        if _has_skip_prefix(status, profile.skip_status_prefixes):
            skipped_prefix += 1
            continue
        if not phone:
            skipped_no_phone += 1
            continue

        seen_tracking.add(tracking)
        status_counter[status or "(empty)"] += 1
        success_flag = 1 if _is_success(status, profile.success_keywords) else 0

        sbds = phone_to_sbds.get(phone)
        if not sbds:
            unmatched_phones.add(phone)
            continue
        for sbd in sbds:
            matched_sbds.add(sbd)
            if success_flag:
                success_count += 1
            to_insert.append(
                (
                    round_id,
                    sbd,
                    tracking,
                    success_flag,
                    status,
                    row.get(col_sent_at, "").strip() if col_sent_at else "",
                    phone,
                    row.get(col_address, "").strip() if col_address else "",
                    row.get(col_recipient, "").strip() if col_recipient else "",
                    carrier_name,
                    synced_at,
                )
            )

    inserted = 0
    if commit:
        with closing(sqlite3.connect(str(db_path))) as conn, conn:
            _ensure_schema(conn)
            conn.executemany(
                "INSERT OR REPLACE INTO shipment_history "
                "(round_id, sbd, tracking_code, is_success, status_raw, "
                "sent_at, phone, address, recipient, carrier, synced_at) "
                "VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)",
                to_insert,
            )
            inserted = len(to_insert)

    stats = BulkImportStats(
        carrier=carrier_name,
        round_id=round_id,
        parsed=len(rows),
        skipped_prefix=skipped_prefix,
        skipped_no_tracking=skipped_no_tracking,
        skipped_no_phone=skipped_no_phone,
        matched_sbds=len(matched_sbds),
        unmatched_phones=len(unmatched_phones),
        inserted=inserted,
        success_count=success_count,
        committed=commit,
        status_breakdown=dict(status_counter),
    )

    log_admin_action(
        activity,
        user_id=admin_user_id,
        user_email=admin_email,
        action="shipment.bulk_import",
        target_id=f"{round_id}:{carrier_name}",
        metadata={
            "parsed": stats.parsed,
            "matched_sbds": stats.matched_sbds,
            "inserted": stats.inserted,
            "success_count": stats.success_count,
            "unmatched_phones": stats.unmatched_phones,
            "committed": stats.committed,
        },
        ip=client_ip,
    )
    return stats


def _iso_now() -> str:
    import time

    return time.strftime("%Y-%m-%dT%H:%M:%SZ", time.gmtime())
