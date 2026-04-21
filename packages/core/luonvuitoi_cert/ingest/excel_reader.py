"""Read students from an Excel workbook (.xlsx / .xlsm).

First row is the header; subsequent rows are data. Cells are coerced to
strings at read time — downstream code doesn't have to defensively cast from
numeric/datetime/formula cell types.

We lean on openpyxl directly (not pandas) to keep the dependency surface
small; pandas is still in the core dependency list for a later phase, but
ingest doesn't need it.
"""

from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from luonvuitoi_cert.ingest.base import IngestError


def _cell_to_str(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def read_excel(path: str | Path, sheet: str | None = None) -> list[dict[str, str]]:
    """Return list of row-dicts keyed by header-row cell values.

    ``sheet`` selects a worksheet by name; defaults to the active sheet.
    Empty rows (all blank cells) are dropped.
    """
    p = Path(path).expanduser().resolve()
    if not p.exists():
        raise IngestError(f"Excel file not found: {p}")
    try:
        wb = load_workbook(p, read_only=True, data_only=True)
    except Exception as e:
        raise IngestError(f"failed to open Excel workbook ({p}): {e}") from e

    # ``load_workbook(read_only=True)`` holds a file handle open until close();
    # ``finally`` guarantees it releases on every path — Windows otherwise locks
    # the file against subsequent writers in the same test.
    try:
        ws = wb[sheet] if sheet else wb.active
        if ws is None:
            raise IngestError(f"workbook has no active worksheet: {p}")

        rows = ws.iter_rows(values_only=True)
        try:
            header_raw = next(rows)
        except StopIteration:
            return []
        headers = [_cell_to_str(h) for h in header_raw]
        if not any(headers):
            raise IngestError(f"Excel header row is empty: {p}")

        out: list[dict[str, str]] = []
        for row in rows:
            values = [_cell_to_str(v) for v in row]
            values += [""] * (len(headers) - len(values))
            if all(v == "" for v in values[: len(headers)]):
                continue
            out.append({h: v for h, v in zip(headers, values[: len(headers)]) if h})
        return out
    finally:
        wb.close()
