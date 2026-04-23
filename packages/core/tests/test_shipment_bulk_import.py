"""Tests for :mod:`luonvuitoi_cert.shipment.bulk_import`."""

from __future__ import annotations

import sqlite3
from pathlib import Path

import pytest
from luonvuitoi_cert.auth import ActivityLog
from luonvuitoi_cert.config import CertConfig
from luonvuitoi_cert.shipment import (
    BulkImportError,
    BulkImportStats,
    bulk_import_shipments,
)

# ---------- fixtures -------------------------------------------------------


@pytest.fixture
def import_config_dict(config_dict: dict) -> dict:
    config_dict["features"] = {
        "shipment": {
            "enabled": True,
            "statuses": ["pending", "shipped", "delivered"],
            "fields": ["tracking_code", "carrier"],
            "import": {
                "default": "viettel",
                "profiles": {
                    "viettel": {
                        "column_mapping": {
                            "tracking_code": ["Mã vận đơn", "Tracking"],
                            "phone": ["SĐT", "Phone"],
                            "status": ["Trạng thái", "Status"],
                            "sent_at": ["Ngày gửi"],
                            "address": ["Địa chỉ"],
                            "recipient": ["Người nhận"],
                        },
                        "success_keywords": ["GIAO THÀNH CÔNG", "DELIVERED"],
                        "skip_status_prefixes": ["CH"],
                        "header_row": 0,
                    },
                    "ghn": {
                        "column_mapping": {
                            "tracking_code": "Order Code",
                            "phone": "Phone",
                            "status": "Status",
                        },
                        "success_keywords": ["DELIVERED"],
                    },
                },
            },
        }
    }
    return config_dict


@pytest.fixture
def import_config(import_config_dict: dict) -> CertConfig:
    return CertConfig.model_validate(import_config_dict)


@pytest.fixture
def import_populated_db(import_config: CertConfig, project_root: Path, tmp_path: Path) -> Path:
    from luonvuitoi_cert.ingest import ingest_rows

    db = tmp_path / "test.db"
    ingest_rows(
        import_config,
        db,
        "main",
        [
            {
                "sbd": "10001",
                "full_name": "Học Viên A",
                "dob": "2010-01-15",
                "school": "A",
                "phone": "0901000001",
                "s": "GOLD",
            },
            {
                "sbd": "10002",
                "full_name": "Học Viên B",
                "dob": "2010-02-20",
                "school": "B",
                "phone": "0901000002",
                "s": "SILVER",
            },
            {
                "sbd": "10003",
                "full_name": "Học Viên C",
                "dob": "2010-03-25",
                "school": "C",
                "phone": "0901000003",
                "s": "GOLD",
            },
        ],
    )
    return db


@pytest.fixture
def carrier_xlsx(tmp_path: Path) -> Path:
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.append(["Mã vận đơn", "SĐT", "Trạng thái", "Ngày gửi", "Địa chỉ", "Người nhận"])
    ws.append(["VN001", "0901000001", "GIAO THÀNH CÔNG", "01/04/2026", "Addr A", "Parent A"])
    ws.append(["VN002", "0901000002", "ĐANG GIAO", "01/04/2026", "Addr B", "Parent B"])
    ws.append(["VN001", "0901000001", "GIAO THÀNH CÔNG", "01/04/2026", "Addr A", "Parent A"])  # dup
    ws.append(["CH-X9", "0901000003", "CHỜ CHUYỂN", "02/04/2026", "Addr C", "Parent C"])  # skip prefix
    ws.append(["VN003", "0909999999", "GIAO THÀNH CÔNG", "01/04/2026", "?", "?"])  # unmatched phone
    ws.append(["VN004", "", "GIAO THÀNH CÔNG", "", "", ""])  # skip no phone
    ws.append(["", "0901000002", "GIAO THÀNH CÔNG", "", "", ""])  # skip no tracking
    path = tmp_path / "viettel-export.xlsx"
    wb.save(path)
    return path


@pytest.fixture
def carrier_csv(tmp_path: Path) -> Path:
    path = tmp_path / "viettel-export.csv"
    path.write_text(
        "\ufeffMã vận đơn,SĐT,Trạng thái\nVN001,0901000001,GIAO THÀNH CÔNG\nVN002,0901000002,ĐANG GIAO\n",
        encoding="utf-8",
    )
    return path


def _audit(tmp_path: Path) -> ActivityLog:
    return ActivityLog(tmp_path / "audit.db")


# ---------- tests ----------------------------------------------------------


def test_rejects_config_without_import_block(
    cert_config, populated_db, tmp_path: Path, carrier_xlsx: Path
) -> None:  # type: ignore[no-untyped-def]
    """config without features.shipment.import must error out."""
    with pytest.raises(BulkImportError, match="shipment.import block missing"):
        bulk_import_shipments(
            config=cert_config,
            db_path=populated_db,
            activity=_audit(tmp_path),
            file_path=carrier_xlsx,
            round_id="main",
        )


def test_rejects_when_phone_col_missing(import_config_dict: dict, tmp_path: Path, carrier_xlsx: Path) -> None:
    from luonvuitoi_cert.ingest import ingest_rows

    import_config_dict["data_mapping"].pop("phone_col")
    cfg = CertConfig.model_validate(import_config_dict)
    db = tmp_path / "test.db"
    ingest_rows(cfg, db, "main", [{"sbd": "10001", "full_name": "X", "s": "GOLD"}])
    with pytest.raises(BulkImportError, match="phone_col must be set"):
        bulk_import_shipments(
            config=cfg, db_path=db, activity=_audit(tmp_path), file_path=carrier_xlsx, round_id="main"
        )


def test_rejects_unknown_carrier(
    import_config, import_populated_db, tmp_path: Path, carrier_xlsx: Path
) -> None:  # type: ignore[no-untyped-def]
    with pytest.raises(BulkImportError, match="unknown carrier"):
        bulk_import_shipments(
            config=import_config,
            db_path=import_populated_db,
            activity=_audit(tmp_path),
            file_path=carrier_xlsx,
            round_id="main",
            carrier="nonexistent",
        )


def test_uses_default_carrier_when_none_given(
    import_config, import_populated_db, tmp_path: Path, carrier_xlsx: Path
) -> None:  # type: ignore[no-untyped-def]
    stats = bulk_import_shipments(
        config=import_config,
        db_path=import_populated_db,
        activity=_audit(tmp_path),
        file_path=carrier_xlsx,
        round_id="main",
    )
    assert stats.carrier == "viettel"


def test_parses_xlsx(import_config, import_populated_db, tmp_path: Path, carrier_xlsx: Path) -> None:  # type: ignore[no-untyped-def]
    stats = bulk_import_shipments(
        config=import_config,
        db_path=import_populated_db,
        activity=_audit(tmp_path),
        file_path=carrier_xlsx,
        round_id="main",
        carrier="viettel",
    )
    assert stats.parsed == 7
    assert stats.skipped_prefix == 1  # CH-X9
    assert stats.skipped_no_tracking == 1
    assert stats.skipped_no_phone == 1
    assert stats.matched_sbds == 2  # 10001, 10002
    assert stats.unmatched_phones == 1  # 0909999999
    assert stats.inserted == 0  # dry-run default
    assert stats.committed is False


def test_parses_csv(import_config, import_populated_db, tmp_path: Path, carrier_csv: Path) -> None:  # type: ignore[no-untyped-def]
    stats = bulk_import_shipments(
        config=import_config,
        db_path=import_populated_db,
        activity=_audit(tmp_path),
        file_path=carrier_csv,
        round_id="main",
        carrier="viettel",
    )
    assert stats.parsed == 2
    assert stats.matched_sbds == 2


def test_dedup_by_tracking(import_config, import_populated_db, tmp_path: Path, carrier_xlsx: Path) -> None:  # type: ignore[no-untyped-def]
    """Duplicate tracking row in input → counted once."""
    stats = bulk_import_shipments(
        config=import_config,
        db_path=import_populated_db,
        activity=_audit(tmp_path),
        file_path=carrier_xlsx,
        round_id="main",
        carrier="viettel",
        commit=True,
    )
    with sqlite3.connect(import_populated_db) as conn:
        count = conn.execute(
            "SELECT COUNT(*) FROM shipment_history WHERE tracking_code = 'VN001'"
        ).fetchone()[0]
    assert count == 1
    assert stats.inserted == 2  # VN001 + VN002, dedup worked


def test_success_keyword_detection(
    import_config, import_populated_db, tmp_path: Path, carrier_xlsx: Path
) -> None:  # type: ignore[no-untyped-def]
    stats = bulk_import_shipments(
        config=import_config,
        db_path=import_populated_db,
        activity=_audit(tmp_path),
        file_path=carrier_xlsx,
        round_id="main",
        carrier="viettel",
        commit=True,
    )
    assert stats.success_count == 1  # only VN001 is GIAO THÀNH CÔNG
    with sqlite3.connect(import_populated_db) as conn:
        rows = conn.execute(
            "SELECT tracking_code, is_success FROM shipment_history ORDER BY tracking_code"
        ).fetchall()
    assert rows == [("VN001", 1), ("VN002", 0)]


def test_append_history_keeps_all_tracking_codes(
    import_config,  # type: ignore[no-untyped-def]
    import_populated_db,
    tmp_path: Path,
) -> None:
    """Same SBD, different tracking codes → multiple history rows."""
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.append(["Mã vận đơn", "SĐT", "Trạng thái"])
    ws.append(["VN001", "0901000001", "ĐANG GIAO"])
    ws.append(["VN002-RESEND", "0901000001", "GIAO THÀNH CÔNG"])
    path = tmp_path / "two-attempts.xlsx"
    wb.save(path)

    bulk_import_shipments(
        config=import_config,
        db_path=import_populated_db,
        activity=_audit(tmp_path),
        file_path=path,
        round_id="main",
        carrier="viettel",
        commit=True,
    )
    with sqlite3.connect(import_populated_db) as conn:
        rows = conn.execute(
            "SELECT tracking_code, is_success FROM shipment_history "
            "WHERE sbd = '10001' ORDER BY tracking_code"
        ).fetchall()
    assert rows == [("VN001", 0), ("VN002-RESEND", 1)]


def test_dry_run_does_not_write_db(
    import_config, import_populated_db, tmp_path: Path, carrier_xlsx: Path
) -> None:  # type: ignore[no-untyped-def]
    bulk_import_shipments(
        config=import_config,
        db_path=import_populated_db,
        activity=_audit(tmp_path),
        file_path=carrier_xlsx,
        round_id="main",
        carrier="viettel",
        commit=False,
    )
    with sqlite3.connect(import_populated_db) as conn:
        tables = [r[0] for r in conn.execute("SELECT name FROM sqlite_master WHERE type='table'")]
    assert "shipment_history" not in tables


def test_idempotent_replace(import_config, import_populated_db, tmp_path: Path, carrier_xlsx: Path) -> None:  # type: ignore[no-untyped-def]
    """Rerun with same file → INSERT OR REPLACE keeps row count stable."""
    for _ in range(2):
        bulk_import_shipments(
            config=import_config,
            db_path=import_populated_db,
            activity=_audit(tmp_path),
            file_path=carrier_xlsx,
            round_id="main",
            carrier="viettel",
            commit=True,
        )
    with sqlite3.connect(import_populated_db) as conn:
        count = conn.execute("SELECT COUNT(*) FROM shipment_history").fetchone()[0]
    assert count == 2  # VN001 + VN002 (history preserved by tracking code)


def test_rejects_unresolved_header(import_config, import_populated_db, tmp_path: Path) -> None:  # type: ignore[no-untyped-def]
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.append(["Wrong Column", "Another", "Thứ ba"])
    ws.append(["x", "y", "z"])
    path = tmp_path / "bad-headers.xlsx"
    wb.save(path)

    with pytest.raises(BulkImportError, match="headers.*unresolved"):
        bulk_import_shipments(
            config=import_config,
            db_path=import_populated_db,
            activity=_audit(tmp_path),
            file_path=path,
            round_id="main",
            carrier="viettel",
        )


def test_missing_file_raises(import_config, import_populated_db, tmp_path: Path) -> None:  # type: ignore[no-untyped-def]
    with pytest.raises(BulkImportError, match="not found"):
        bulk_import_shipments(
            config=import_config,
            db_path=import_populated_db,
            activity=_audit(tmp_path),
            file_path=tmp_path / "nonexistent.xlsx",
            round_id="main",
            carrier="viettel",
        )


def test_unsupported_file_type(import_config, import_populated_db, tmp_path: Path) -> None:  # type: ignore[no-untyped-def]
    p = tmp_path / "file.txt"
    p.write_text("not a spreadsheet")
    with pytest.raises(BulkImportError, match="unsupported file type"):
        bulk_import_shipments(
            config=import_config,
            db_path=import_populated_db,
            activity=_audit(tmp_path),
            file_path=p,
            round_id="main",
            carrier="viettel",
        )


def test_logs_bulk_import_audit(
    import_config, import_populated_db, tmp_path: Path, carrier_xlsx: Path
) -> None:  # type: ignore[no-untyped-def]
    log = _audit(tmp_path)
    bulk_import_shipments(
        config=import_config,
        db_path=import_populated_db,
        activity=log,
        file_path=carrier_xlsx,
        round_id="main",
        carrier="viettel",
        admin_user_id="u1",
        admin_email="admin@test.co",
        commit=True,
    )
    entries = log.recent()
    assert entries[0].action == "shipment.bulk_import"
    assert entries[0].target_id == "main:viettel"
    assert entries[0].metadata["parsed"] == 7
    assert entries[0].metadata["committed"] is True


def test_header_row_offset_xlsx(import_config_dict: dict, tmp_path: Path) -> None:
    """Metadata rows above the real header (common in carrier exports)."""
    from luonvuitoi_cert.ingest import ingest_rows
    from openpyxl import Workbook

    import_config_dict["features"]["shipment"]["import"]["profiles"]["viettel"]["header_row"] = 2
    cfg = CertConfig.model_validate(import_config_dict)

    db = tmp_path / "test.db"
    ingest_rows(cfg, db, "main", [{"sbd": "10001", "full_name": "X", "phone": "0901000001", "s": "GOLD"}])

    wb = Workbook()
    ws = wb.active
    ws.append(["Carrier Monthly Export", "", ""])
    ws.append(["Generated:", "2026-04-01", ""])
    ws.append(["Mã vận đơn", "SĐT", "Trạng thái"])  # row 3 = index 2
    ws.append(["VN001", "0901000001", "GIAO THÀNH CÔNG"])
    path = tmp_path / "with-header-offset.xlsx"
    wb.save(path)

    stats = bulk_import_shipments(
        config=cfg,
        db_path=db,
        activity=_audit(tmp_path),
        file_path=path,
        round_id="main",
        carrier="viettel",
        commit=True,
    )
    assert stats.parsed == 1
    assert stats.matched_sbds == 1
    assert stats.success_count == 1


def test_stats_frozen_dataclass() -> None:
    from dataclasses import FrozenInstanceError

    stats = BulkImportStats(carrier="viettel", round_id="main")
    with pytest.raises(FrozenInstanceError):
        stats.inserted = 1  # type: ignore[misc]


def test_phone_normalized_handles_float_and_separators(
    import_config,  # type: ignore[no-untyped-def]
    import_populated_db,
    tmp_path: Path,
) -> None:
    """Phone stored as float ('901000001.0') or with spaces ('0901 000 001')
    should still normalize to the same digits as '0901000001' and match.
    """
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.append(["Mã vận đơn", "SĐT", "Trạng thái"])
    ws.append(["VN001", 901000001, "DELIVERED"])  # numeric cell — pandas/openpyxl stores as number
    ws.append(["VN002", "0901 000 002", "DELIVERED"])  # spaces
    path = tmp_path / "messy-phone.xlsx"
    wb.save(path)

    stats = bulk_import_shipments(
        config=import_config,
        db_path=import_populated_db,
        activity=_audit(tmp_path),
        file_path=path,
        round_id="main",
        carrier="viettel",
    )
    assert stats.matched_sbds == 2
