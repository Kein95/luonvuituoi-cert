"""Tests for the shipment-draft state machine (`luonvuitoi_cert.shipment.draft`)."""

from __future__ import annotations

import io
import sqlite3
from pathlib import Path

import pytest
from luonvuitoi_cert.auth import ActivityLog, Role, issue_admin_token
from luonvuitoi_cert.config import CertConfig
from luonvuitoi_cert.shipment import (
    DraftError,
    draft_add,
    draft_cancel,
    draft_export,
    draft_list,
)
from luonvuitoi_cert.shipment.bulk_import import bulk_import_shipments

# ---------- fixtures -------------------------------------------------------


@pytest.fixture
def draft_config_dict(config_dict: dict) -> dict:
    config_dict["features"] = {
        "shipment": {
            "enabled": True,
            "statuses": ["pending", "shipped", "delivered"],
            "fields": ["tracking_code", "carrier"],
            "import": {
                "default": "viettel",
                "profiles": {
                    "viettel": {
                        "column_mapping": {
                            "tracking_code": ["Mã vận đơn"],
                            "phone": ["SĐT"],
                            "status": ["Trạng thái"],
                        },
                        "success_keywords": ["GIAO THÀNH CÔNG"],
                        "export_template": {
                            "sbd": "Mã học viên",
                            "full_name": "Họ tên",
                            "phone": "SĐT",
                            "address": "Địa chỉ nhận",
                            "recipient": "Người nhận",
                        },
                    }
                },
            },
        }
    }
    # Add ship_method extra col for filter tests
    config_dict["data_mapping"]["extra_cols"] = ["ship_method"]
    return config_dict


@pytest.fixture
def draft_config(draft_config_dict: dict) -> CertConfig:
    return CertConfig.model_validate(draft_config_dict)


@pytest.fixture
def draft_populated_db(draft_config: CertConfig, project_root: Path, tmp_path: Path) -> Path:
    from luonvuitoi_cert.ingest import ingest_rows

    db = tmp_path / "test.db"
    ingest_rows(
        draft_config,
        db,
        "main",
        [
            {
                "sbd": "10001",
                "full_name": "Student A",
                "dob": "2010-01-15",
                "school": "School A",
                "phone": "0901000001",
                "ship_method": "CA_NHAN",
                "s": "GOLD",
            },
            {
                "sbd": "10002",
                "full_name": "Student B",
                "dob": "2010-02-20",
                "school": "School B",
                "phone": "0901000002",
                "ship_method": "CA_NHAN",
                "s": "SILVER",
            },
            {
                "sbd": "10003",
                "full_name": "Student C",
                "dob": "2010-03-25",
                "school": "School C",
                "phone": "0901000003",
                "ship_method": "TRUONG",
                "s": "GOLD",
            },
        ],
    )
    return db


def _env() -> dict[str, str]:
    return {"JWT_SECRET": "test-secret-please-change-32-bytes-min"}


def _admin_token(role: Role = Role.ADMIN) -> str:
    return issue_admin_token(user_id="u1", email="a@b.co", role=role, env=_env())


# ---------- tests ----------------------------------------------------------


def test_add_requires_filter(draft_config, draft_populated_db, tmp_path: Path) -> None:  # type: ignore[no-untyped-def]
    log = ActivityLog(tmp_path / "audit.db")
    with pytest.raises(DraftError, match="at least one of"):
        draft_add(
            config=draft_config,
            db_path=draft_populated_db,
            activity=log,
            params={"token": _admin_token(), "round_id": "main"},
            env=_env(),
        )


def test_add_rejects_viewer(draft_config, draft_populated_db, tmp_path: Path) -> None:  # type: ignore[no-untyped-def]
    log = ActivityLog(tmp_path / "audit.db")
    with pytest.raises(DraftError, match="read-only|cannot manage"):
        draft_add(
            config=draft_config,
            db_path=draft_populated_db,
            activity=log,
            params={
                "token": _admin_token(Role.VIEWER),
                "round_id": "main",
                "filters": {"ship_method": "CA_NHAN"},
            },
            env=_env(),
        )


def test_add_by_column_filter(draft_config, draft_populated_db, tmp_path: Path) -> None:  # type: ignore[no-untyped-def]
    log = ActivityLog(tmp_path / "audit.db")
    created = draft_add(
        config=draft_config,
        db_path=draft_populated_db,
        activity=log,
        params={
            "token": _admin_token(),
            "round_id": "main",
            "filters": {"ship_method": "CA_NHAN"},
        },
        env=_env(),
    )
    assert {d.sbd for d in created} == {"10001", "10002"}
    assert all(d.status == "draft" for d in created)


def test_add_by_result_filter(draft_config, draft_populated_db, tmp_path: Path) -> None:  # type: ignore[no-untyped-def]
    log = ActivityLog(tmp_path / "audit.db")
    created = draft_add(
        config=draft_config,
        db_path=draft_populated_db,
        activity=log,
        params={"token": _admin_token(), "round_id": "main", "result": "GOLD"},
        env=_env(),
    )
    assert {d.sbd for d in created} == {"10001", "10003"}


def test_add_by_sbd_list(draft_config, draft_populated_db, tmp_path: Path) -> None:  # type: ignore[no-untyped-def]
    log = ActivityLog(tmp_path / "audit.db")
    created = draft_add(
        config=draft_config,
        db_path=draft_populated_db,
        activity=log,
        params={
            "token": _admin_token(),
            "round_id": "main",
            "sbd_list": ["10001", "10003"],
        },
        env=_env(),
    )
    assert {d.sbd for d in created} == {"10001", "10003"}


def test_add_idempotent_skips_existing_active_draft(draft_config, draft_populated_db, tmp_path: Path) -> None:  # type: ignore[no-untyped-def]
    log = ActivityLog(tmp_path / "audit.db")
    first = draft_add(
        config=draft_config,
        db_path=draft_populated_db,
        activity=log,
        params={
            "token": _admin_token(),
            "round_id": "main",
            "filters": {"ship_method": "CA_NHAN"},
        },
        env=_env(),
    )
    second = draft_add(
        config=draft_config,
        db_path=draft_populated_db,
        activity=log,
        params={
            "token": _admin_token(),
            "round_id": "main",
            "filters": {"ship_method": "CA_NHAN"},
        },
        env=_env(),
    )
    assert len(first) == 2
    assert second == []  # both already exist as 'draft'


def test_add_rejects_unknown_filter_column(draft_config, draft_populated_db, tmp_path: Path) -> None:  # type: ignore[no-untyped-def]
    log = ActivityLog(tmp_path / "audit.db")
    with pytest.raises(DraftError, match="not in students schema"):
        draft_add(
            config=draft_config,
            db_path=draft_populated_db,
            activity=log,
            params={
                "token": _admin_token(),
                "round_id": "main",
                "filters": {"not_a_column": "whatever"},
            },
            env=_env(),
        )


def test_list_by_round_and_status(draft_config, draft_populated_db, tmp_path: Path) -> None:  # type: ignore[no-untyped-def]
    log = ActivityLog(tmp_path / "audit.db")
    draft_add(
        config=draft_config,
        db_path=draft_populated_db,
        activity=log,
        params={
            "token": _admin_token(),
            "round_id": "main",
            "filters": {"ship_method": "CA_NHAN"},
        },
        env=_env(),
    )
    listed = draft_list(
        config=draft_config,
        db_path=draft_populated_db,
        params={"token": _admin_token(), "round_id": "main", "status": "draft"},
        env=_env(),
    )
    assert len(listed) == 2


def test_cancel_affects_only_draft_and_exported(draft_config, draft_populated_db, tmp_path: Path) -> None:  # type: ignore[no-untyped-def]
    log = ActivityLog(tmp_path / "audit.db")
    created = draft_add(
        config=draft_config,
        db_path=draft_populated_db,
        activity=log,
        params={
            "token": _admin_token(),
            "round_id": "main",
            "filters": {"ship_method": "CA_NHAN"},
        },
        env=_env(),
    )
    affected = draft_cancel(
        config=draft_config,
        db_path=draft_populated_db,
        activity=log,
        params={"token": _admin_token(), "ids": [d.id for d in created]},
        env=_env(),
    )
    assert affected == 2
    remaining = draft_list(
        config=draft_config,
        db_path=draft_populated_db,
        params={"token": _admin_token(), "status": "cancelled"},
        env=_env(),
    )
    assert len(remaining) == 2


def test_export_produces_xlsx_and_locks_drafts(draft_config, draft_populated_db, tmp_path: Path) -> None:  # type: ignore[no-untyped-def]
    log = ActivityLog(tmp_path / "audit.db")
    draft_add(
        config=draft_config,
        db_path=draft_populated_db,
        activity=log,
        params={
            "token": _admin_token(),
            "round_id": "main",
            "filters": {"ship_method": "CA_NHAN"},
        },
        env=_env(),
    )
    result = draft_export(
        config=draft_config,
        db_path=draft_populated_db,
        activity=log,
        params={"token": _admin_token(), "round_id": "main", "carrier": "viettel"},
        env=_env(),
    )
    assert result.row_count == 2
    assert result.filename.endswith(".xlsx")
    assert result.file_bytes[:2] == b"PK"  # xlsx is a zip

    # Drafts now locked as 'exported'
    listed = draft_list(
        config=draft_config,
        db_path=draft_populated_db,
        params={"token": _admin_token(), "status": "exported"},
        env=_env(),
    )
    assert len(listed) == 2
    assert all(d.batch_id == result.batch_id for d in listed)


def test_export_no_drafts_raises(draft_config, draft_populated_db, tmp_path: Path) -> None:  # type: ignore[no-untyped-def]
    log = ActivityLog(tmp_path / "audit.db")
    with pytest.raises(DraftError, match="no 'draft'-status rows"):
        draft_export(
            config=draft_config,
            db_path=draft_populated_db,
            activity=log,
            params={"token": _admin_token(), "round_id": "main", "carrier": "viettel"},
            env=_env(),
        )


def test_export_unknown_carrier_raises(draft_config, draft_populated_db, tmp_path: Path) -> None:  # type: ignore[no-untyped-def]
    log = ActivityLog(tmp_path / "audit.db")
    with pytest.raises(DraftError, match="unknown carrier"):
        draft_export(
            config=draft_config,
            db_path=draft_populated_db,
            activity=log,
            params={"token": _admin_token(), "round_id": "main", "carrier": "ghosted"},
            env=_env(),
        )


def test_exported_draft_excel_has_template_headers(draft_config, draft_populated_db, tmp_path: Path) -> None:  # type: ignore[no-untyped-def]
    from openpyxl import load_workbook

    log = ActivityLog(tmp_path / "audit.db")
    draft_add(
        config=draft_config,
        db_path=draft_populated_db,
        activity=log,
        params={
            "token": _admin_token(),
            "round_id": "main",
            "sbd_list": ["10001"],
        },
        env=_env(),
    )
    result = draft_export(
        config=draft_config,
        db_path=draft_populated_db,
        activity=log,
        params={"token": _admin_token(), "round_id": "main", "carrier": "viettel"},
        env=_env(),
    )
    wb = load_workbook(io.BytesIO(result.file_bytes))
    ws = wb.active
    headers = [c.value for c in next(ws.iter_rows())]
    assert headers == ["Mã học viên", "Họ tên", "SĐT", "Địa chỉ nhận", "Người nhận"]
    data = [c.value for c in list(ws.iter_rows())[1]]
    assert data[0] == "10001"  # SBD
    assert data[2] == "0901000001"  # phone
    assert data[3] in (None, "")  # address column blank (admin fills later)


def test_bulk_import_promotes_exported_drafts(draft_config, draft_populated_db, tmp_path: Path) -> None:  # type: ignore[no-untyped-def]
    from openpyxl import Workbook

    log = ActivityLog(tmp_path / "audit.db")
    draft_add(
        config=draft_config,
        db_path=draft_populated_db,
        activity=log,
        params={
            "token": _admin_token(),
            "round_id": "main",
            "sbd_list": ["10001"],
        },
        env=_env(),
    )
    draft_export(
        config=draft_config,
        db_path=draft_populated_db,
        activity=log,
        params={"token": _admin_token(), "round_id": "main", "carrier": "viettel"},
        env=_env(),
    )

    # Simulate carrier return — tracking file
    wb = Workbook()
    ws = wb.active
    ws.append(["Mã vận đơn", "SĐT", "Trạng thái"])
    ws.append(["VN-AAA-001", "0901000001", "GIAO THÀNH CÔNG"])
    carrier_file = tmp_path / "carrier-return.xlsx"
    wb.save(carrier_file)

    bulk_import_shipments(
        config=draft_config,
        db_path=draft_populated_db,
        activity=log,
        file_path=carrier_file,
        round_id="main",
        carrier="viettel",
        commit=True,
    )

    with sqlite3.connect(draft_populated_db) as conn:
        conn.row_factory = sqlite3.Row
        row = conn.execute(
            "SELECT status, tracking_code, promoted_at FROM shipment_draft WHERE sbd='10001'"
        ).fetchone()
    assert row["status"] == "promoted"
    assert row["tracking_code"] == "VN-AAA-001"
    assert row["promoted_at"] is not None


def test_add_logs_audit(draft_config, draft_populated_db, tmp_path: Path) -> None:  # type: ignore[no-untyped-def]
    log = ActivityLog(tmp_path / "audit.db")
    draft_add(
        config=draft_config,
        db_path=draft_populated_db,
        activity=log,
        params={
            "token": _admin_token(),
            "round_id": "main",
            "filters": {"ship_method": "CA_NHAN"},
        },
        env=_env(),
    )
    entries = log.recent()
    assert entries[0].action == "shipment.draft.add"
    assert entries[0].metadata["created"] == 2
