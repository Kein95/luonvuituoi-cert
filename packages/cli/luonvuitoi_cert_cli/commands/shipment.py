"""``lvt-cert shipment <subcommand>`` — manage shipment drafts + exports.

Subcommands:
  - draft add     Add drafts from a student filter (column / result / SBD list).
  - draft list    Show drafts (optionally filtered by round / status / batch).
  - draft cancel  Void one or more drafts by ID.
  - export        Produce carrier-ready Excel and hard-lock drafts as 'exported'.

All subcommands require JWT_SECRET (env) and an admin token; pass via
``--token`` or ``LVT_ADMIN_TOKEN`` env. Handlers are pure — this module is a
thin Typer shell around them.
"""

from __future__ import annotations

import json as _json
import os
from pathlib import Path

import typer
from rich.console import Console
from rich.table import Table

console = Console()
app = typer.Typer(help="Manage shipment drafts + exports.", no_args_is_help=True)
draft_app = typer.Typer(help="Draft lifecycle: add / list / cancel.", no_args_is_help=True)
app.add_typer(draft_app, name="draft")


def _load_token(explicit: str | None) -> str:
    tok = explicit or os.getenv("LVT_ADMIN_TOKEN", "").strip()
    if not tok:
        console.print("[red]ERR[/] admin token required — pass --token or set LVT_ADMIN_TOKEN")
        raise typer.Exit(code=2)
    return tok


def _load_ctx(config_path: Path, db_path: Path | None):  # type: ignore[no-untyped-def]
    from luonvuitoi_cert.auth import ActivityLog
    from luonvuitoi_cert.config import load_config

    try:
        config = load_config(config_path)
    except Exception as e:  # noqa: BLE001
        console.print(f"[red]ERR[/] can't load config {config_path}: {e}")
        raise typer.Exit(code=2) from e
    resolved_db = (db_path or Path("data") / f"{config.project.slug}.db").expanduser().resolve()
    if not resolved_db.exists():
        console.print(f"[red]ERR[/] DB not found: {resolved_db}")
        raise typer.Exit(code=2)
    activity = ActivityLog(resolved_db, gsheet_webhook_url=os.getenv("GSHEET_WEBHOOK_URL"))
    return config, resolved_db, activity


@draft_app.command("add")
def draft_add_cmd(
    round_id: str = typer.Option("main", "--round", "-r"),
    filter_expr: list[str] = typer.Option([], "--filter", "-f", help="column=value filter (repeatable)."),
    result: str | None = typer.Option(None, "--result", help="Match any subject column holding this result."),
    sbd_file: Path | None = typer.Option(None, "--from-file", help="Excel/CSV with SBDs to draft."),
    note: str | None = typer.Option(None, "--note"),
    token: str | None = typer.Option(None, "--token"),
    config_path: Path = typer.Option(Path("cert.config.json"), "--config", "-c"),
    db_path: Path | None = typer.Option(None, "--db"),
) -> None:
    from luonvuitoi_cert.shipment import DraftError, draft_add

    config, db, activity = _load_ctx(config_path, db_path)
    sbd_list: list[str] | None = None
    if sbd_file is not None:
        sbd_list = _read_sbd_list(sbd_file)
    params: dict = {
        "token": _load_token(token),
        "round_id": round_id,
        "filters": filter_expr or None,
        "result": result,
        "sbd_list": sbd_list,
        "note": note,
    }
    try:
        created = draft_add(config=config, db_path=db, activity=activity, params=params)
    except DraftError as e:
        console.print(f"[red]ERR[/] {e}")
        raise typer.Exit(code=1) from e
    console.print(f"[green]OK[/] created [cyan]{len(created)}[/] draft(s)")


@draft_app.command("list")
def draft_list_cmd(
    round_id: str | None = typer.Option(None, "--round", "-r"),
    status: str | None = typer.Option(None, "--status"),
    batch_id: str | None = typer.Option(None, "--batch"),
    limit: int = typer.Option(100, "--limit"),
    json_output: bool = typer.Option(False, "--json"),
    token: str | None = typer.Option(None, "--token"),
    config_path: Path = typer.Option(Path("cert.config.json"), "--config", "-c"),
    db_path: Path | None = typer.Option(None, "--db"),
) -> None:
    from luonvuitoi_cert.shipment import DraftError, draft_list

    config, db, _ = _load_ctx(config_path, db_path)
    params: dict = {
        "token": _load_token(token),
        "round_id": round_id,
        "status": status,
        "batch_id": batch_id,
        "limit": limit,
    }
    try:
        rows = draft_list(config=config, db_path=db, params=params)
    except DraftError as e:
        console.print(f"[red]ERR[/] {e}")
        raise typer.Exit(code=1) from e

    if json_output:
        from dataclasses import asdict

        print(_json.dumps([asdict(r) for r in rows], ensure_ascii=False, indent=2))
        return

    if not rows:
        console.print("[dim]no drafts matched[/]")
        return
    table = Table("Status", "Round", "SBD", "Carrier", "Batch", "Updated")
    for r in rows:
        table.add_row(r.status, r.round_id, r.sbd, r.carrier or "—", (r.batch_id or "—")[:8], r.updated_at)
    console.print(table)


@draft_app.command("cancel")
def draft_cancel_cmd(
    ids: list[str] = typer.Argument(..., help="Draft IDs to cancel."),
    token: str | None = typer.Option(None, "--token"),
    config_path: Path = typer.Option(Path("cert.config.json"), "--config", "-c"),
    db_path: Path | None = typer.Option(None, "--db"),
) -> None:
    from luonvuitoi_cert.shipment import DraftError, draft_cancel

    config, db, activity = _load_ctx(config_path, db_path)
    try:
        n = draft_cancel(
            config=config,
            db_path=db,
            activity=activity,
            params={"token": _load_token(token), "ids": ids},
        )
    except DraftError as e:
        console.print(f"[red]ERR[/] {e}")
        raise typer.Exit(code=1) from e
    console.print(f"[green]OK[/] cancelled [cyan]{n}[/] draft(s)")


@app.command("export")
def export_cmd(
    round_id: str = typer.Option("main", "--round", "-r"),
    carrier: str = typer.Option(..., "--carrier"),
    output: Path | None = typer.Option(
        None, "--output", "-o", help="Output file path (default: batch-id-based)."
    ),
    token: str | None = typer.Option(None, "--token"),
    config_path: Path = typer.Option(Path("cert.config.json"), "--config", "-c"),
    db_path: Path | None = typer.Option(None, "--db"),
) -> None:
    from luonvuitoi_cert.shipment import DraftError, draft_export

    config, db, activity = _load_ctx(config_path, db_path)
    try:
        result = draft_export(
            config=config,
            db_path=db,
            activity=activity,
            params={"token": _load_token(token), "round_id": round_id, "carrier": carrier},
        )
    except DraftError as e:
        console.print(f"[red]ERR[/] {e}")
        raise typer.Exit(code=1) from e
    out_path = (output or Path(result.filename)).expanduser().resolve()
    out_path.parent.mkdir(parents=True, exist_ok=True)
    out_path.write_bytes(result.file_bytes)
    console.print(
        f"[green]EXPORTED[/] {result.row_count} row(s) → [cyan]{out_path}[/]\n"
        f"  batch: {result.batch_id}  (drafts now hard-locked)"
    )


def _read_sbd_list(path: Path) -> list[str]:
    from openpyxl import load_workbook

    suffix = path.suffix.lower()
    if suffix in {".xlsx", ".xlsm"}:
        wb = load_workbook(path, read_only=True, data_only=True)
        try:
            ws = wb.active
            out: list[str] = []
            for row in ws.iter_rows(values_only=True):
                for cell in row:
                    s = str(cell or "").strip()
                    if s:
                        out.append(s)
                        break
            return out
        finally:
            wb.close()
    if suffix == ".csv":
        import csv

        with path.open(encoding="utf-8-sig", newline="") as f:
            return [row[0].strip() for row in csv.reader(f) if row and row[0].strip()]
    raise typer.BadParameter(f"unsupported suffix {suffix!r}; use .xlsx/.xlsm/.csv")
