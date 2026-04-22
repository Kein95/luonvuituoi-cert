"""Integration tests for the Phase 11 CLI commands: init, seed, dev app builder."""

from __future__ import annotations

import json
from pathlib import Path

import pytest
from luonvuitoi_cert_cli.main import app
from typer.testing import CliRunner

runner = CliRunner()


# ── init ────────────────────────────────────────────────────────────


def test_init_scaffolds_new_project(tmp_path: Path) -> None:
    target = tmp_path / "my-portal"
    result = runner.invoke(
        app,
        [
            "init",
            str(target),
            "--name",
            "Demo Academy",
            "--slug",
            "demo-academy",
            "--locale",
            "en",
            "--non-interactive",
        ],
    )
    assert result.exit_code == 0, result.stdout
    assert (target / "cert.config.json").exists()
    assert (target / "vercel.json").exists()
    assert (target / "requirements.txt").exists()
    assert (target / "README.md").exists()
    # Phase 15: scaffold ships the Vercel entrypoint.
    assert (target / "api" / "index.py").exists()


def test_scaffolded_vercel_entrypoint_builds_flask_app(tmp_path: Path) -> None:
    """Regression: the scaffolded api/index.py must load a working Flask app."""
    target = tmp_path / "v-portal"
    runner.invoke(app, ["init", str(target), "--slug", "v-portal", "--non-interactive"])
    # Scaffold the remaining assets the Flask build_app needs (templates + a
    # placeholder PDF + a real font so the config validates at runtime).
    (target / "templates").mkdir(exist_ok=True)
    import shutil

    import reportlab
    from reportlab.pdfgen import canvas

    vera = Path(reportlab.__file__).parent / "fonts" / "Vera.ttf"
    (target / "assets" / "fonts").mkdir(parents=True, exist_ok=True)
    shutil.copy2(vera, target / "assets" / "fonts" / "serif.ttf")
    pdf = target / "templates" / "main.pdf"
    c = canvas.Canvas(str(pdf), pagesize=(842, 595))
    c.drawString(100, 500, "T")
    c.showPage()
    c.save()

    # Import the scaffolded entrypoint as a module, assert the app exists.
    import importlib.util

    spec = importlib.util.spec_from_file_location("_scaffold_entry", target / "api" / "index.py")
    module = importlib.util.module_from_spec(spec)  # type: ignore[arg-type]
    spec.loader.exec_module(module)  # type: ignore[union-attr]
    assert hasattr(module, "app")
    client = module.app.test_client()
    resp = client.get("/")
    assert resp.status_code == 200


def test_vercel_json_rewrites_catch_all(tmp_path: Path) -> None:
    """Regression: Phase 15 consolidates routes under /api/index."""
    import json as _json

    target = tmp_path / "vv"
    runner.invoke(app, ["init", str(target), "--slug", "vv-portal", "--non-interactive"])
    data = _json.loads((target / "vercel.json").read_text(encoding="utf-8"))
    rewrites = data["rewrites"]
    assert len(rewrites) == 1
    assert rewrites[0]["source"] == "/(.*)"
    assert rewrites[0]["destination"] == "/api/index"


def test_init_refuses_non_empty_target(tmp_path: Path) -> None:
    (tmp_path / "existing.txt").write_text("hi", encoding="utf-8")
    result = runner.invoke(app, ["init", str(tmp_path), "--non-interactive"])
    assert result.exit_code != 0


def test_init_rejects_invalid_slug(tmp_path: Path) -> None:
    target = tmp_path / "x"
    result = runner.invoke(
        app,
        ["init", str(target), "--slug", "Not_A_Slug!", "--non-interactive"],
    )
    assert result.exit_code != 0


def test_init_rejects_invalid_locale(tmp_path: Path) -> None:
    target = tmp_path / "x"
    result = runner.invoke(
        app,
        ["init", str(target), "--slug", "demo", "--locale", "xx", "--non-interactive"],
    )
    assert result.exit_code != 0


def test_init_rendered_config_validates(tmp_path: Path) -> None:
    target = tmp_path / "v"
    result = runner.invoke(
        app,
        ["init", str(target), "--name", "V", "--slug", "v-demo", "--non-interactive"],
    )
    assert result.exit_code == 0
    data = json.loads((target / "cert.config.json").read_text(encoding="utf-8"))
    assert data["project"]["name"] == "V"
    assert data["project"]["slug"] == "v-demo"


# ── seed ────────────────────────────────────────────────────────────


def test_seed_generates_excel(tmp_path: Path) -> None:
    target = tmp_path / "seed-test"
    runner.invoke(
        app,
        ["init", str(target), "--slug", "seed-test", "--non-interactive"],
    )
    output = target / "data" / "students.xlsx"
    result = runner.invoke(
        app,
        [
            "seed",
            "--count",
            "4",
            "--seed",
            "7",
            "--output",
            str(output),
            "--config",
            str(target / "cert.config.json"),
        ],
    )
    assert result.exit_code == 0, result.stdout
    assert output.exists()

    from openpyxl import load_workbook

    wb = load_workbook(output)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    assert len(rows) == 5  # header + 4 students
    wb.close()


def test_seed_deterministic_with_seed_value(tmp_path: Path) -> None:
    target = tmp_path / "d"
    runner.invoke(app, ["init", str(target), "--slug", "d", "--non-interactive"])
    out1 = target / "data" / "a.xlsx"
    out2 = target / "data" / "b.xlsx"
    for out in (out1, out2):
        runner.invoke(
            app,
            [
                "seed",
                "--count",
                "3",
                "--seed",
                "12345",
                "--output",
                str(out),
                "--config",
                str(target / "cert.config.json"),
            ],
        )
    assert out1.read_bytes() != b""
    # Same seed => same students => same XLSX body (modulo openpyxl's timestamps).
    # Compare the rendered rows instead of raw bytes.
    from openpyxl import load_workbook

    rows_a = list(load_workbook(out1).active.iter_rows(values_only=True))
    rows_b = list(load_workbook(out2).active.iter_rows(values_only=True))
    assert rows_a == rows_b


# ── dev app ─────────────────────────────────────────────────────────


@pytest.fixture
def scaffolded_project(tmp_path: Path) -> Path:
    target = tmp_path / "live"
    runner.invoke(
        app,
        ["init", str(target), "--slug", "live-portal", "--non-interactive"],
    )
    # Seed so the DB exists with a student.
    runner.invoke(
        app,
        [
            "seed",
            "--count",
            "2",
            "--seed",
            "1",
            "--output",
            str(target / "data" / "students.xlsx"),
            "--config",
            str(target / "cert.config.json"),
        ],
    )
    return target


def test_build_app_registers_pages_and_api(scaffolded_project: Path) -> None:
    from luonvuitoi_cert_cli.server import build_app

    flask_app = build_app(scaffolded_project / "cert.config.json", scaffolded_project)
    rules = {r.rule for r in flask_app.url_map.iter_rules()}
    assert "/" in rules
    assert "/admin" in rules
    assert "/certificate-checker" in rules
    assert "/api/search" in rules
    assert "/api/download" in rules
    assert "/api/verify" in rules
    assert "/api/captcha" in rules
    assert "/api/admin/login" in rules


def test_dev_server_serves_portal_html(scaffolded_project: Path) -> None:
    from luonvuitoi_cert_cli.server import build_app

    flask_app = build_app(scaffolded_project / "cert.config.json", scaffolded_project)
    client = flask_app.test_client()
    resp = client.get("/")
    assert resp.status_code == 200
    assert b"<!doctype html>" in resp.data.lower()
    assert b"Certificate Portal" in resp.data or b"Cert" in resp.data


def test_dev_server_admin_emits_csp(scaffolded_project: Path) -> None:
    """/admin must carry a Content-Security-Policy header to protect the sessionStorage JWT."""
    from luonvuitoi_cert_cli.server import build_app

    flask_app = build_app(scaffolded_project / "cert.config.json", scaffolded_project)
    client = flask_app.test_client()
    resp = client.get("/admin")
    assert resp.status_code == 200
    assert "Content-Security-Policy" in resp.headers


def test_dev_server_captcha_endpoint(scaffolded_project: Path) -> None:
    from luonvuitoi_cert_cli.server import build_app

    flask_app = build_app(scaffolded_project / "cert.config.json", scaffolded_project)
    client = flask_app.test_client()
    resp = client.post("/api/captcha")
    assert resp.status_code == 200
    body = resp.get_json()
    assert "id" in body and "question" in body


def test_flask_enforces_max_content_length(scaffolded_project: Path) -> None:
    """Regression: Phase 11 review H1 — Werkzeug must reject oversize bodies pre-parse."""
    from luonvuitoi_cert_cli.server import build_app

    flask_app = build_app(scaffolded_project / "cert.config.json", scaffolded_project)
    client = flask_app.test_client()
    resp = client.post("/api/search", data=b"x" * (40 * 1024), content_type="application/json")
    assert resp.status_code == 413


def test_csp_uses_nonce_not_unsafe_inline(scaffolded_project: Path) -> None:
    """Regression: Phase 11 review H4 — CSP script-src must not contain unsafe-inline."""
    import re

    from luonvuitoi_cert_cli.server import build_app

    flask_app = build_app(scaffolded_project / "cert.config.json", scaffolded_project)
    client = flask_app.test_client()
    resp = client.get("/admin")
    csp = resp.headers.get("Content-Security-Policy", "")
    script_section = csp.split("script-src", 1)[1].split(";")[0] if "script-src" in csp else ""
    assert "'unsafe-inline'" not in script_section
    assert "'nonce-" in csp
    nonce_match = re.search(r"'nonce-([^']+)'", csp)
    assert nonce_match
    assert f'nonce="{nonce_match.group(1)}"' in resp.data.decode("utf-8")


def test_login_errorhandler_narrows_to_login_error(scaffolded_project: Path) -> None:
    """Regression: Phase 11 review H2 — internal exceptions must not leak as 401 body."""
    from luonvuitoi_cert_cli.server import build_app

    flask_app = build_app(scaffolded_project / "cert.config.json", scaffolded_project)
    client = flask_app.test_client()
    resp = client.post("/api/admin/login", json={})
    assert resp.status_code == 401
    assert resp.get_json()["error"] == "email and password are required"


def test_public_base_url_env_overrides_host(
    scaffolded_project: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    """Regression: Phase 11 review H5 — PUBLIC_BASE_URL must win over Host header."""
    from luonvuitoi_cert_cli.server import build_app
    from luonvuitoi_cert_cli.server.app import _public_base_url

    flask_app = build_app(scaffolded_project / "cert.config.json", scaffolded_project)
    monkeypatch.setenv("PUBLIC_BASE_URL", "https://trusted.example")
    with flask_app.test_request_context("/", headers={"Host": "evil.example"}):
        assert _public_base_url() == "https://trusted.example"


def test_init_template_escapes_quotes_in_project_name(tmp_path: Path) -> None:
    """Regression: Phase 11 review M1 — project name with quotes must not break JSON."""
    target = tmp_path / "escape-dir"
    result = runner.invoke(
        app,
        [
            "init",
            str(target),
            "--name",
            'Evil"; "injected":"yes',
            "--slug",
            "evil-demo",
            "--non-interactive",
        ],
    )
    assert result.exit_code == 0, result.stdout
    data = json.loads((target / "cert.config.json").read_text(encoding="utf-8"))
    assert data["project"]["name"] == 'Evil"; "injected":"yes'


def test_init_leaves_target_clean_on_validation_failure(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    """Regression: Phase 11 review M2 — failed validation must not leave a half-populated target."""
    import luonvuitoi_cert.config as cfg_module
    from luonvuitoi_cert.config.loader import ConfigError

    target = tmp_path / "atomic"

    def _always_fail(path):  # type: ignore[no-untyped-def]
        raise ConfigError("forced failure")

    monkeypatch.setattr(cfg_module, "load_config", _always_fail)
    result = runner.invoke(
        app,
        ["init", str(target), "--slug", "atomic-demo", "--non-interactive"],
    )
    assert result.exit_code != 0
    assert not target.exists()
